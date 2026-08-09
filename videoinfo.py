#!/usr/bin/env python3
"""批量将 B 站视频 TAG 名称和 music_id 写入 Excel 文件。

脚本会从 A 列读取 BVID,先查询每个视频的 CID,再调用支持 BGM 信息的
B 站 TAG 接口。TAG 名称会写入原工作簿的 R 列,music_id 会写入 S 列。

用法:
    python bilibili_video_tags_to_excel.py videos.xlsx
    python bilibili_video_tags_to_excel.py videos.xlsx --cookie "SESSDATA=xxx"
    python bilibili_video_tags_to_excel.py videos.xlsx --sheet Sheet1 --start-row 2

依赖:
    pip install openpyxl requests
"""

from __future__ import annotations

import argparse
import shutil
import sys
import time
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

VIEW_URL = "https://api.bilibili.com/x/web-interface/view"
TAG_URL = "https://api.bilibili.com/x/web-interface/view/detail/tag"

BVID_COLUMN = "A"
TAG_NAME_COLUMN = "R"
MUSIC_ID_COLUMN = "S"


class BilibiliAPIError(RuntimeError):
    """B 站接口返回错误时抛出。"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从 Excel 的 A 列读取 BVID,并把 tag_name/music_id 写入 R/S 列。"
    )
    parser.add_argument("excel", type=Path, help="需要原地更新的 .xlsx 文件路径。")
    parser.add_argument("--sheet", help="工作表名称。默认使用当前活动工作表。")
    parser.add_argument("--start-row", type=int, default=2, help="第一行数据所在行号。默认:2。")
    parser.add_argument(
        "--separator", default="; ", help="单个视频存在多个 TAG 时使用的分隔符。默认:'; '。"
    )
    parser.add_argument(
        "--cookie",
        default="",
        help="可选的 Cookie 请求头,例如 'SESSDATA=xxx'。参考接口可能需要登录态。",
    )
    parser.add_argument("--timeout", type=float, default=10.0, help="HTTP 超时时间,单位秒。默认:10。")
    parser.add_argument("--sleep", type=float, default=0.3, help="每个视频之间的请求间隔,单位秒。默认:0.3。")
    parser.add_argument(
        "--no-backup", action="store_true", help="覆盖工作簿前不创建 .bak.xlsx 备份。"
    )
    return parser.parse_args()


def api_get(session: requests.Session, url: str, params: dict[str, Any], timeout: float) -> dict[str, Any]:
    response = session.get(url, params=params, timeout=timeout)
    response.raise_for_status()
    payload = response.json()
    if payload.get("code") != 0:
        raise BilibiliAPIError(f"{url} 返回 code={payload.get('code')}, message={payload.get('message')}")
    return payload


def get_first_cid(session: requests.Session, bvid: str, timeout: float) -> int:
    payload = api_get(session, VIEW_URL, {"bvid": bvid}, timeout)
    pages = payload.get("data", {}).get("pages") or []
    if not pages or not pages[0].get("cid"):
        raise BilibiliAPIError(f"未查询到 cid: {bvid}")
    return int(pages[0]["cid"])


def get_tags(session: requests.Session, bvid: str, cid: int, timeout: float) -> list[dict[str, Any]]:
    payload = api_get(session, TAG_URL, {"bvid": bvid, "cid": cid}, timeout)
    data = payload.get("data")
    if not isinstance(data, list):
        raise BilibiliAPIError(f"TAG 接口返回格式异常: {bvid}: data 不是列表")
    return data


def make_session(cookie: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/126.0 Safari/537.36",
            "Referer": "https://www.bilibili.com/",
        }
    )
    if cookie:
        session.headers["Cookie"] = cookie
    return session


def main() -> int:
    args = parse_args()
    if not args.excel.exists():
        print(f"未找到 Excel 文件: {args.excel}", file=sys.stderr)
        return 2

    workbook = load_workbook(args.excel)
    worksheet = workbook[args.sheet] if args.sheet else workbook.active
    session = make_session(args.cookie)

    processed = 0
    failed: list[tuple[int, str, str]] = []

    for row in range(args.start_row, worksheet.max_row + 1):
        raw_bvid = worksheet[f"{BVID_COLUMN}{row}"].value
        bvid = str(raw_bvid).strip() if raw_bvid is not None else ""
        if not bvid:
            continue

        try:
            cid = get_first_cid(session, bvid, args.timeout)
            tags = get_tags(session, bvid, cid, args.timeout)
            tag_names = [str(tag.get("tag_name", "")).strip() for tag in tags if tag.get("tag_name")]
            music_ids = [str(tag.get("music_id", "")).strip() for tag in tags if tag.get("music_id")]
            worksheet[f"{TAG_NAME_COLUMN}{row}"] = args.separator.join(tag_names)
            worksheet[f"{MUSIC_ID_COLUMN}{row}"] = args.separator.join(music_ids)
            processed += 1
            print(f"row {row}: {bvid} -> {len(tag_names)} 个 TAG, {len(music_ids)} 个 music_id")
        except Exception as exc:  # noqa: BLE001 - 单行失败后继续批量处理。
            failed.append((row, bvid, str(exc)))
            print(f"row {row}: {bvid} 失败: {exc}", file=sys.stderr)

        if args.sleep > 0:
            time.sleep(args.sleep)

    if not args.no_backup:
        backup_path = args.excel.with_suffix(args.excel.suffix + ".bak.xlsx")
        shutil.copy2(args.excel, backup_path)
        print(f"已写入备份: {backup_path}")

    workbook.save(args.excel)
    print(f"已更新 {processed} 行,文件: {args.excel}")

    if failed:
        print("失败明细:", file=sys.stderr)
        for row, bvid, reason in failed:
            print(f"  row {row} ({bvid}): {reason}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

